"""
Deterministic Excel Content Ingestion Tool
Reads MIROR_Content_Library.xlsx without mutating the source file and converts operational Excel rows into structured renderer payloads.
"""

import os
import sys
import json
from pathlib import Path
from typing import Dict, List, Any, Tuple

# Ensure UTF-8 console output encoding
sys.stdout.reconfigure(encoding="utf-8")

REPO_ROOT = Path(__file__).resolve().parents[2]
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

try:
    import openpyxl
except ImportError:
    raise RuntimeError("openpyxl dependency is required for Excel ingestion.")

def parse_excel_content(excel_path: Path) -> List[Dict[str, Any]]:
    """
    Parses operational content rows from Excel workbook without mutating the source file.
    Returns a list of structured post payload dictionaries.
    """
    if not excel_path.exists():
        raise FileNotFoundError(f"Excel workbook not found at: {excel_path}")

    # Load workbook read-only to guarantee zero file mutation
    wb = openpyxl.load_workbook(excel_path, data_only=True, read_only=True)
    if "Content_Library" not in wb.sheetnames:
        raise ValueError("Workbook is missing required sheet 'Content_Library'.")

    ws = wb["Content_Library"]
    
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("Sheet 'Content_Library' is completely empty.")

    header = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
    
    # Map header column names to indexes
    col_map = {}
    for idx, name in enumerate(header):
        if name:
            col_map[name] = idx

    required_cols = ["Content ID", "SLIDE 1 — HOOK", "SLIDE 2 — FOLLOW-THROUGH", "SLIDE 3 — CTA"]
    for req in required_cols:
        if req not in col_map:
            # Fallback check for ascii hyphens if em dashes differ
            matched = [k for k in col_map if req.split(' ')[0] in k]
            if not matched:
                raise ValueError(f"Missing required column '{req}' in Content_Library header.")

    parsed_posts = []
    seen_ids = set()

    for r_idx, row in enumerate(rows[1:], start=2):
        if not any(cell is not None for cell in row):
            continue

        cid = str(row[col_map["Content ID"]]).strip() if row[col_map["Content ID"]] is not None else ""
        if not cid:
            continue

        if cid in seen_ids:
            raise ValueError(f"Duplicate Content ID '{cid}' detected at Excel row {r_idx}.")
        seen_ids.add(cid)

        status_val = str(row[col_map.get("Status", 11)]).strip() if col_map.get("Status") and row[col_map["Status"]] is not None else "READY"
        title_val = str(row[col_map.get("Content Title", 1)]).strip() if col_map.get("Content Title") and row[col_map["Content Title"]] is not None else cid

        s1_raw = str(row[col_map["SLIDE 1 — HOOK"]]).strip() if row[col_map["SLIDE 1 — HOOK"]] is not None else ""
        s2_raw = str(row[col_map["SLIDE 2 — FOLLOW-THROUGH"]]).strip() if row[col_map["SLIDE 2 — FOLLOW-THROUGH"]] is not None else ""
        s3_raw = str(row[col_map["SLIDE 3 — CTA"]]).strip() if row[col_map["SLIDE 3 — CTA"]] is not None else ""

        if not s1_raw or not s2_raw or not s3_raw:
            raise ValueError(f"Content item '{cid}' at row {r_idx} is missing required slide text.")

        # Decompose Slide 1
        s1_headline = s1_raw

        # Decompose Slide 2
        s2_parts = [p.strip() for p in s2_raw.split("\n\n") if p.strip()]
        if not s2_parts:
            s2_parts = [p.strip() for p in s2_raw.split("\n") if p.strip()]
        s2_headline = s2_parts[0]
        s2_body_texts = s2_parts[1:] if len(s2_parts) > 1 else []

        # Decompose Slide 3
        s3_parts = [p.strip() for p in s3_raw.split("\n\n") if p.strip()]
        s3_headline = s3_parts[0]
        
        cta_text = None
        s3_body_raw_blocks = []

        if len(s3_parts) > 1 and ("JOIN THE MIROR" in s3_parts[-1] or "Link in bio" in s3_parts[-1] or "→" in s3_parts[-1]):
            cta_text = s3_parts[-1]
            s3_body_raw_blocks = s3_parts[1:-1]
        else:
            s3_body_raw_blocks = s3_parts[1:]
            cta_text = "JOIN THE MIROR COMMUNITY →\nLink in bio"

        s3_body_texts = []
        for block in s3_body_raw_blocks:
            lines = [line.strip() for line in block.split("\n") if line.strip()]
            s3_body_texts.extend(lines)

        post_payload = {
            "post_id": cid,
            "content_id": cid,
            "template": "T01",
            "template_id": "T01",
            "content_title": title_val,
            "status": status_val,
            "source": "MIROR_Content_Library.xlsx",
            "slides": [
                {
                    "id": "S01",
                    "slide": "S01",
                    "type": "hook",
                    "headline": {"text": s1_headline, "lock": "EXACT"}
                },
                {
                    "id": "S02",
                    "slide": "S02",
                    "type": "follow-through",
                    "headline": {"text": s2_headline, "lock": "EXACT"},
                    "body": [{"text": b, "lock": "EXACT"} for b in s2_body_texts]
                },
                {
                    "id": "S03",
                    "slide": "S03",
                    "type": "cta",
                    "headline": {"text": s3_headline, "lock": "EXACT"},
                    "body": [{"text": b, "lock": "EXACT"} for b in s3_body_texts],
                    "cta": {"text": cta_text, "lock": "EXACT"}
                }
            ]
        }
        parsed_posts.append(post_payload)

    wb.close()
    return parsed_posts

if __name__ == "__main__":
    excel_file = Path(r"C:\Users\Hp\Downloads\MIROR_Content_Library.xlsx")
    try:
        posts = parse_excel_content(excel_file)
        print(f"=== INGESTION SUCCESS: Successfully parsed {len(posts)} operational content items from Excel ===")
        for p in posts:
            print(f"  Post ID: {p['post_id']} | Title: {p['content_title']} | Status: {p['status']}")
    except Exception as e:
        print(f"=== INGESTION ERROR: {str(e)} ===")
        sys.exit(1)
